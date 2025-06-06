import hashlib
import os
import zipfile
import tempfile
import shutil
import logging
import multiprocessing
from enum import Enum
from functools import partial
from importlib.resources import files
from typing import List, Dict, Tuple, Optional
from collections import Counter

import arpy
from tqdm import tqdm

import numpy as np
import pandas as pd
import tensorflow as tf
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from elftools.elf.elffile import ELFFile

# File analysis status mapping
FILE_STATUS_MAPPING = {
    'analyzed': 'Successfully Analyzed',
    'skipped': 'Skipped (System Library)',
    'failed': 'Analysis Failed',
}


class FileType(Enum):
    SO = 1
    AR = 2
    NOT_SUPPORT = 0xff


class FileInfo:
    """Represents information about a binary file"""
    TEXT_SECTION = '.text'

    def __init__(self, absolute_path: str, logical_path: Optional[str] = None):
        self.absolute_path = absolute_path
        self.logical_path = logical_path or absolute_path
        self.file_size = self._get_file_size()
        self.file_hash = self._calculate_file_hash()
        self.file_id = self._generate_file_id()
        if absolute_path.endswith('.a'):
            self.file_type = FileType.AR
        elif absolute_path.endswith('.so'):
            self.file_type = FileType.SO
        else:
            self.file_type = FileType.NOT_SUPPORT

    def __repr__(self) -> str:
        return f"FileInfo({self.logical_path}, size={self.file_size}, hash={self.file_hash[:8]}...)"

    def to_dict(self) -> dict:
        return {
            'absolute_path': self.absolute_path,
            'logical_path': self.logical_path,
            'file_size': self.file_size,
            'file_hash': self.file_hash,
            'file_id': self.file_id
        }

    def extract_dot_text(self) -> List[int]:
        """Extract .text segment data"""
        if self.file_type == FileType.SO:
            return self._extract_so_dot_text(self.absolute_path)
        elif self.file_type == FileType.AR:
            return self._extract_archive_dot_text()
        return []

    def _extract_so_dot_text(self, file_path) -> List[int]:
        try:
            with open(file_path, 'rb') as f:
                elf = ELFFile(f)
                section = elf.get_section_by_name(self.TEXT_SECTION)
                if section:
                    return list(section.data())
        except Exception as e:
            logging.error("Failed to extract .text section from %s: %s", file_path, e)
        return []

    def _extract_archive_dot_text(self) -> List[int]:
        text_data = []
        try:
            ar = arpy.Archive(self.absolute_path)
            for name in ar.namelist():
                elf = ELFFile(ar.open(name))
                section = elf.get_section_by_name(self.TEXT_SECTION)
                if section:
                    text_data.extend(list(section.data()))
        except Exception as e:
            logging.error("Failed to extract archive file %s: %s", self.absolute_path, e)
        return text_data

    def _get_file_size(self) -> int:
        return os.path.getsize(self.absolute_path)

    def _calculate_file_hash(self) -> str:
        with open(self.absolute_path, 'rb') as f:
            return hashlib.md5(f.read()).hexdigest()

    def _generate_file_id(self) -> str:
        base_name = os.path.basename(self.absolute_path).replace(' ', '_')
        unique_id = f"{base_name}_{self.file_hash}"
        return self.file_hash if len(unique_id) > 200 else unique_id


class OptimizationDetector:
    MATCHED_FILES_DIR = 'matched_files_results'

    def __init__(self, workers: int = 1):
        self.parallel = workers > 1
        self.workers = min(workers, multiprocessing.cpu_count() - 1)
        self.model = None
        self.temp_dirs = []

    @staticmethod
    def _merge_chunk_results(df: pd.DataFrame) -> Dict[str, dict]:
        results = {}
        for file, group in df.groupby('file'):
            predictions = group['prediction'].tolist()
            distribution = dict(Counter(predictions))
            most_common = Counter(predictions).most_common(1)[0][0]
            confidence = predictions.count(most_common) / len(predictions)
            total_chunks = len(predictions)

            if any(x in distribution for x in [0, 1, 2, 3, 4]):
                opt_score = (
                                    distribution.get(0, 0) * 0 +  # O0
                                    distribution.get(1, 0) * 0.33 +  # O1
                                    distribution.get(2, 0) * 0.67 +  # O2
                                    distribution.get(3, 0) * 1.0 +  # O3
                                    distribution.get(4, 0) * 0.67  # Os
                            ) / total_chunks

                # Determine optimization category
                if opt_score < 0.2:
                    opt_category = "Unoptimized (O0 dominant)"
                elif opt_score < 0.4:
                    opt_category = "Low Optimization (O1 dominant)"
                elif opt_score < 0.7:
                    opt_category = "Medium Optimization (O2/Os dominant)"
                else:
                    opt_category = "High Optimization (O3 dominant)"
            else:
                opt_score = None
                opt_category = None

            results[file] = {
                'prediction': most_common,
                'confidence': confidence,
                'distribution': distribution,
                'opt_score': opt_score,
                'opt_category': opt_category,
                'total_chunks': total_chunks
            }

        return results

    def detect_optimization(self, input_path: str, output: str = "binary_analysis_report.xlsx"):
        file_infos = self._collect_binary_files(input_path)
        if not file_infos:
            logging.warning("No valid binary files found")
            return
        success, failures = self._analyze_files(file_infos, output)
        self.cleanup()
        logging.info("Analysis complete: %s files analyzed, %s files failed", success, failures)

    def cleanup(self):
        for temp_dir in self.temp_dirs:
            shutil.rmtree(temp_dir, ignore_errors=True)
        self.temp_dirs = []
        self.model = None

    def _extract_hap_file(self, hap_path: str) -> List[FileInfo]:
        """Extract SO files from HAP/HSP archives and return FileInfo objects"""
        extracted_files = []
        temp_dir = tempfile.mkdtemp()
        self.temp_dirs.append(temp_dir)

        try:
            with zipfile.ZipFile(hap_path, 'r') as zip_ref:
                for file in zip_ref.namelist():
                    if file.startswith('libs/arm64') and file.endswith('.so'):
                        output_path = os.path.join(temp_dir, file[5:])
                        os.makedirs(os.path.dirname(output_path), exist_ok=True)
                        with zip_ref.open(file) as src, open(output_path, 'wb') as dest:
                            dest.write(src.read())
                        file_info = FileInfo(
                            absolute_path=output_path,
                            logical_path=f"{hap_path}/{file}"
                        )
                        extracted_files.append(file_info)
        except Exception as e:
            logging.error("Failed to extract HAP file %s: %s", hap_path, e)
        return extracted_files

    def _collect_binary_files(self, input_path: str) -> List[FileInfo]:
        """Collect binary files for analysis"""
        file_infos = []
        if os.path.isfile(input_path):
            if input_path.endswith(('.so', '.a')):
                file_infos.append(FileInfo(input_path))
            elif input_path.endswith(('.hap', '.hsp')):
                file_infos.extend(self._extract_hap_file(input_path))
        elif os.path.isdir(input_path):
            for root, _, _files in os.walk(input_path):
                for file in _files:
                    file_path = os.path.join(root, file)
                    if file.endswith(('.so', '.a')):
                        logical_path = os.path.relpath(file_path, input_path)
                        file_infos.append(FileInfo(file_path, logical_path))
                    elif file.endswith(('.hap', '.hsp')):
                        file_infos.extend(self._extract_hap_file(file_path))
        return file_infos

    @staticmethod
    def _extract_features(file_info: FileInfo, features: int = 2048) -> Optional[np.ndarray]:
        data = file_info.extract_dot_text()
        if not data or len(data) == 0:
            return None

        sequences = []
        while data:
            seq = data[:features]
            if len(seq) < features:
                seq = np.pad(seq, (0, features - len(seq)), 'constant')
            sequences.append(seq)
            data = data[features:]
        return np.array(sequences, dtype=np.uint8)

    def _run_inference(self, file_info: FileInfo, model, features: int = 2048) -> List[Tuple[int, float]]:
        features_array = self._extract_features(file_info, features)
        if features_array is None or not features_array.size:
            return []

        y_predict = model.predict(features_array, batch_size=256)
        results = []
        for _predict in y_predict:
            prediction = np.argmax(_predict)
            confidence = _predict[prediction]
            results.append((prediction, confidence))
        return results

    def _run_analysis(self, file_info: FileInfo) -> Tuple[FileInfo, Optional[List[Tuple[int, float]]]]:
        """Run optimization flag detection on a single file"""
        # Lazy load model
        if self.model is None:
            flags_model = files('hapray.optimization_detector').joinpath("models/aarch64-flag-lstm-converted.h5")
            self.model = tf.keras.models.load_model(str(flags_model))

        return file_info, self._run_inference(file_info, self.model)

    def _analyze_files(self, file_infos: List[FileInfo], output_file: str) -> Tuple[int, int]:
        # Filter out analyzed files
        remaining_files = []
        for file_info in file_infos:
            flags_path = os.path.join(self.MATCHED_FILES_DIR, f"flags_{file_info.file_id}.csv")
            if os.path.exists(flags_path):
                logging.debug("Skipping already analyzed file: %s", file_info.absolute_path)
                continue
            remaining_files.append(file_info)

        logging.info("Files to analyze: %d", len(remaining_files))

        # Create directory for intermediate results
        os.makedirs(self.MATCHED_FILES_DIR, exist_ok=True)

        if remaining_files:
            process_func = partial(self._run_analysis)
            if self.parallel and len(remaining_files) > 1:
                logging.info("Using %d parallel workers", self.workers)
                with multiprocessing.Pool(self.workers) as pool:
                    results = list(tqdm(
                        pool.imap(process_func, remaining_files),
                        total=len(remaining_files),
                        desc="Analyzing binaries"
                    ))
            else:
                results = [process_func(fi) for fi in tqdm(remaining_files, desc="Analyzing binaries")]

            # Save intermediate results
            for file_info, flags in results:
                if flags:
                    flags_path = os.path.join(self.MATCHED_FILES_DIR, f"flags_{file_info.file_id}.csv")
                    with open(flags_path, "w", encoding='UTF-8') as f:
                        f.write("file,prediction,confidence\n")
                        for pred, conf in flags:
                            f.write(f"{file_info.file_id},{pred},{conf}\n")

        flags_results = {}
        files_with_results = 0
        for file_info in file_infos:
            flags_path = os.path.join(self.MATCHED_FILES_DIR, f"flags_{file_info.file_id}.csv")
            if os.path.exists(flags_path):
                try:
                    flags_df = pd.read_csv(flags_path)
                    if not flags_df.empty:
                        file_results = self._merge_chunk_results(flags_df)
                        flags_results.update(file_results)
                        files_with_results += 1
                except Exception as e:
                    logging.error("Error loading results for %s: %s", file_info.absolute_path, e)

        if file_infos:
            self._generate_excel_report(flags_results, file_infos, output_file)
        return files_with_results, len(file_infos) - files_with_results

    def _generate_excel_report(self, flags_results: dict, file_infos: List[FileInfo],
                               output_file: str) -> None:
        wb = Workbook()
        summary_sheet = wb.active
        summary_sheet.title = "Summary"

        headers = [
            "Binary File", "Status", "Optimization Category",
            "Optimization Score", "O0 Chunks", "O1 Chunks", "O2 Chunks",
            "O3 Chunks", "Os Chunks", "Total Chunks", "File Size (bytes)",
            "Size Optimized", "Notes"
        ]

        for col, header in enumerate(headers, 1):
            summary_sheet.cell(row=1, column=col, value=header)

        row = 2
        for file_info in sorted(file_infos, key=lambda x: x.logical_path):
            flags_result = flags_results.get(file_info.file_id)
            if flags_result is None:
                status = FILE_STATUS_MAPPING['failed']
                opt_category = 'N/A'
                opt_score = 'N/A'
                distribution = {0: 0, 1: 0, 2: 0, 3: 0, 4: 0}
                total_chunks = 0
                size_optimized = 'N/A'
            else:
                status = FILE_STATUS_MAPPING['analyzed']
                opt_category = flags_result['opt_category']
                opt_score = flags_result['opt_score']
                distribution = flags_result['distribution']
                total_chunks = flags_result['total_chunks']
                os_chunks = distribution.get(4, 0)
                os_ratio = os_chunks / total_chunks if total_chunks > 0 else 0
                size_optimized = f"{'Yes' if os_ratio >= 0.5 else 'No'} ({os_ratio:.1%})"

            summary_sheet.cell(row=row, column=1, value=file_info.logical_path)
            summary_sheet.cell(row=row, column=2, value=status)
            summary_sheet.cell(row=row, column=3, value=opt_category or "N/A")
            summary_sheet.cell(row=row, column=4,
                               value=f"{opt_score:.2%}" if isinstance(opt_score, float) else opt_score)
            summary_sheet.cell(row=row, column=5, value=distribution.get(0, 0))
            summary_sheet.cell(row=row, column=6, value=distribution.get(1, 0))
            summary_sheet.cell(row=row, column=7, value=distribution.get(2, 0))
            summary_sheet.cell(row=row, column=8, value=distribution.get(3, 0))
            summary_sheet.cell(row=row, column=9, value=distribution.get(4, 0))
            summary_sheet.cell(row=row, column=10, value=total_chunks)
            summary_sheet.cell(row=row, column=11, value=file_info.file_size)
            summary_sheet.cell(row=row, column=12, value=size_optimized)

            # Color-code size optimization status
            if size_optimized.startswith("Yes"):
                summary_sheet.cell(row=row, column=12).fill = PatternFill(
                    start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

            row += 1

        wb.save(output_file)
        logging.info("Report saved to %s", output_file)
